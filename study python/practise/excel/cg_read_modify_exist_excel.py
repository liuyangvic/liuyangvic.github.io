# -*- coding: utf-8 -*-
"""
Created on Wed Mar 17 19:58:21 2021

@author: 182304
"""

import xlrd # read excel lib
import openpyxl # modify exist excel lib

#read excel info
read_file = '01_イニシアティブ_本部内業界活動.xlsx'
read_sheet1_name1 = "タスク・運営委員リスト_202011"

#write excel info
write_file = 'pm_project_jp_listOK.xlsx'
writeh_sheet1_name = "Page 1"

row_data = []

"""
auth_account_pair_2 = []
auth_account_pair_3 = []
"""

def open_read_excel():
        try:
                read_excel = xlrd.open_workbook(filename=read_file) # open excel file
                print("This excel includes following sheets:",read_excel.sheet_names()) #print including sheets name
                return read_excel

        except Exception as e:
                print("FAILED open Excel:",e)
                return -1
            
def open_write_excel():
        try:
                write_excel = openpyxl.load_workbook(filename=write_file) # open excel file
                print("This excel includes following sheets:",write_excel.active.title) #print including sheets name
                write_excel.active = 1 # sheet index #1st from 0
                write_sheet = write_excel.active
                """
                for cell in list(write_sheet.columns)[1]:
                    print(cell.value)
                """    
                for cell in list(write_sheet.rows)[0]:
                    if(cell.value == "承認希望期限"):
                        cell.value =  "承認希望期限_modified"
                    print(cell.value)
                """
                print("before:" ,write_sheet.cell(column=0,row=0).value)
                #write_sheet  = write_excel.worksheets[0]
                write_sheet.cell(0,row=0).value = "afterxxx"
                print("before:" ,write_sheet.cell(column=0,row=0).value)
                """
                #print("This excel includes following sheets:",write_excel.sheet_names()) #print including sheets name
                write_excel.save(filename=write_file)
                return 0

        except Exception as e:
                print("FAILED open Excel:",e)
                return -1

def close_write_excel(write_excel):
        try:
                return write_excel.save(filename=write_file) # open excel file

        except Exception as e:
                print("FAILED open Excel:",e)
                return -1
            
def read_1_sheet(wb,wa):
               
        ###　variable parameters when sheet modified ###　
        sheet1 = wb.sheet_by_name(read_sheet1_name1) # specify sheet by sheet name
        print("=============Sheet %s============="%read_sheet1_name1)
        begin_row = 0 #
        end_row = 3 #
        begin_column = 0 #
        end_column = 3 #
        ###　variable parameters when sheet modified ###　
                
        print("Reading sheet:",sheet1.name, "has" ,sheet1.nrows, "rows" ,sheet1.ncols, "colums" )
           
        for i in range(begin_row,end_row):
                for j in range(begin_column,end_column):
                        row_data.append
                        print(sheet1.cell(i,j).value)
                        
                        """
                        auth_value = sheet1.cell(i,begin_column-1).value
                        account_value = sheet1.cell(begin_row-1,j).value
                        limit_value = sheet1.cell(i,j).value
                        print(auth_value,",",account_value,",",limit_value)

                        # transform limit value        
                        if limit_value == '◎':
                                limit_value = '2'
                        elif limit_value == '○':
                                limit_value = '1'
                        elif limit_value =='×':
                                limit_value = '0'
                        else:
                                print('illegal limit value has read:'+limit_value)
                                limit_value = ''

                        auth_account_pair_1.append([auth_value,account_value,limit_value])                        
                        """                    
                   
def main():
        
        read_excel = open_read_excel()
        write_excel = open_write_excel()

        if read_excel != -1:
                read_1_sheet(read_excel,write_excel)
                """
                read_2_sheet(wb)
                read_3_sheet(wb)
                """
                
        #close_write_excel(write_excel)

main()
