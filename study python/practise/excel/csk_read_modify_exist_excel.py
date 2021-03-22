# -*- coding: utf-8 -*-
"""
Created on Wed Mar 17 19:58:21 2021

@author: 182304
"""

import xlrd # read excel lib
import openpyxl # modify exist excel lib

#read excel info
read_file = '01_イニシアティブ_本部内業界活動.xlsx'
read_sheet1_name1 = "タスク・運営委員リスト_202011"
max_rows = 1

#write excel info
write_file = 'pm_project_jp_listOK.xlsx'
writeh_sheet1_name = "Page 1"

sample_data = []


def open_read_excel():
        try:
                read_excel = xlrd.open_workbook(filename=read_file) # open excel file
                #print("Imported excel includes following sheets:",read_excel.sheet_names()) #print including sheets name
                return read_excel

        except Exception as e:
                print("FAILED open Excel:",e)
                return -1
            
def read_1_sheet(wb):
               
        ###　variable parameters when sheet modified ###　
        sheet1 = wb.sheet_by_name(read_sheet1_name1) # specify sheet by sheet name
        print("=============Read sample data START=============")
        begin_row = 1 #
        #end_row = 10 #
        begin_column = 0 #
        #end_column = 10 #
        ###　variable parameters when sheet modified ###　
                
        print("Imported sheet :",sheet1.name, "has" ,sheet1.nrows, "rows," ,sheet1.ncols, "columns" )
        global max_rows
        max_rows = sheet1.nrows
        #print(max_rows)
        
        for i in range(begin_row,sheet1.nrows):
                for j in range(begin_column,sheet1.ncols):
                        #if sheet1.cell(i,j).value != "":
                            sample_data.append([sheet1.cell(0,j).value,sheet1.cell(i,j).value,i])
                 
        print("=============Read sample data END=============")
            
def open_write_excel():
        try:
                print("=============Write sample data START=============")
                write_excel = openpyxl.load_workbook(filename=write_file) # open excel file
                write_excel.active = 1 # sheet index from 0
                write_sheet = write_excel.active
                print("Write sample data to following sheets:",write_excel.active.title,"has(before)", write_sheet.max_row, "rows,", write_sheet.max_column, "columns") 
                print("--------------")
                #begin_column = 1 # openpyxl index starts from 1
                #begin_row = 2 # openpyxl index starts from 1 # write import data from Row 2
                
                #print(write_sheet.min_column,write_sheet.max_column,write_sheet.min_row,write_sheet.max_row)
                cnt = 1
                for j in range(write_sheet.min_column,write_sheet.max_column+1):
                    if  write_sheet.cell(1,j).value == "終了実績日時"  :
                        for i in sample_data:                            
                            if i[0]== "終了日" :
                                #print(i[0],",",i[1])
                                write_sheet.cell(cnt+1,j).value = i[1]
                                cnt=cnt+1
                print("終了実績日時 write completed!")

                cnt = 1
                for j in range(write_sheet.min_column,write_sheet.max_column+1):
                    if  write_sheet.cell(1,j).value == "終了予定日時" :
                        for i in sample_data:                            
                            if i[0]== "終了日" :
                                #print(i[0],",",i[1])
                                write_sheet.cell(cnt+1,j).value = i[1]
                                cnt=cnt+1
                print("終了予定日時 write completed!")
                
                cnt = 1
                for j in range(write_sheet.min_column,write_sheet.max_column+1):
                    if  write_sheet.cell(1,j).value == "開始実績日時" :
                        for i in sample_data:                            
                            if i[0]== "開始日" :
                                #print(i[0],",",i[1])
                                write_sheet.cell(cnt+1,j).value = i[1]
                                cnt=cnt+1
                print("開始実績日時 write completed!")

                cnt = 1
                for j in range(write_sheet.min_column,write_sheet.max_column+1):
                    if  write_sheet.cell(1,j).value == "開始予定" :
                        for i in sample_data:                            
                            if i[0]== "開始日" :
                                #print(i[0],",",i[1])
                                write_sheet.cell(cnt+1,j).value = i[1]
                                cnt=cnt+1
                print("開始予定 write completed!")
                
                cnt = 1
                for j in range(write_sheet.min_column,write_sheet.max_column+1):
                    if  write_sheet.cell(1,j).value == "プロジェクトメンバ" :
                        for i in sample_data:                            
                            if i[0]== "担当者" :
                                #print(i[0],",",i[1])
                                write_sheet.cell(cnt+1,j).value = i[1]
                                cnt=cnt+1
                print("プロジェクトメンバ write completed!")
                
                cnt = 1
                global max_rows
                for j in range(write_sheet.min_column,write_sheet.max_column+1):
                    if  write_sheet.cell(1,j).value == "管理レベル" :
                        for i in sample_data: 
                            if cnt < max_rows :
                                #print(i[0],",",i[1])
                                write_sheet.cell(cnt+1,j).value = "部署"
                                cnt=cnt+1
                print("管理レベル write completed!")
                
                cnt = 1
                #global max_rows
                for j in range(write_sheet.min_column,write_sheet.max_column+1):
                    if  write_sheet.cell(1,j).value == "ステータス" :
                        for i in sample_data:
                            if cnt < max_rows :                            
                                #print(i[0],",",i[1])
                                write_sheet.cell(cnt+1,j).value = "OPEN"
                                cnt=cnt+1
                print("ステータス write completed!")
                
                cnt = 1
                for j in range(write_sheet.min_column,write_sheet.max_column+1):
                    if  write_sheet.cell(1,j).value == "プロジェクト名" :
                        for i in sample_data:                            
                            if i[0]== "委員・イニシアティブ" :
                                #print(i[0],",",i[1])
                                write_sheet.cell(cnt+1,j).value = i[1]
                                cnt=cnt+1
                print("プロジェクト名 write completed!")
                
                cnt = 1
                #global max_rows
                for j in range(write_sheet.min_column,write_sheet.max_column+1):
                    if  write_sheet.cell(1,j).value == "部門名" :
                        for i in sample_data:
                            #print(max_rows)                            
                            if i[0]== "部署" :
                                #print(i[0],",",i[1])
                                write_sheet.cell(cnt+1,j).value = i[1]
                                cnt=cnt+1
                print("部門名 write completed!")
                
                cnt = 1
                #print(max_rows)
                for j in range(write_sheet.min_column,write_sheet.max_column+1):
                    if  write_sheet.cell(1,j).value == "本部・ユニット名" :
                        for i in sample_data:
                            if cnt < max_rows :
                                #print(i[0],",",i[1])
                                write_sheet.cell(cnt+1,j).value = "臨床開発本部"
                                cnt=cnt+1
                print("本部・ユニット名 write completed!")
                
                #cnt = 1
                combo = []
                x=0
                y=0
                z=0
                for j in range(write_sheet.min_column,write_sheet.max_column+1):
                    if  write_sheet.cell(1,j).value == "説明" :
                        for i in sample_data:                                              
                            if i[0]== "業界団体URL" :
                                #print(i[1])
                                combo.insert(x,"・業界団体URL<br/>"+str(i[1])+"<br/>")
                                #print(combo[x])
                                #write_sheet.cell(x+2,j).value = combo[x]
                                x=x+1
                                #cnt=cnt+1
                                continue
                                                           
                            if i[0]== "委員・イニシアティブ名称" :
                                #print(i[1])
                                combo.insert(y,str(combo[y])+"・委員・イニシアティブ名称<br/>"+str(i[1])+"<br/>")
                                #print(combo[y])
                                y=y+1
                                continue
                                
                            if i[0]== "成果物" :
                                #print(i[1])
                                combo.insert(z,str(combo[z])+"成果物<br/>"+str(i[1])+"<br/>")
                                #print(combo[z])
                                write_sheet.cell(z+2,j).value = combo[z]
                                z=z+1
                                continue
                print("説明 write completed!")

                #cnt = 1
                combo = []
                x=0
                y=0
                z=0
                for j in range(write_sheet.min_column,write_sheet.max_column+1):
                    if  write_sheet.cell(1,j).value == "タグ" :
                        for i in sample_data:                                              
                            if i[0]== "業界団体名" :
                                #print(i[1])
                                #combo.insert(x,"・業界団体名<br/>"+str(i[1])+"<br/>")
                                combo.insert(x,str(i[1])+"<br/>")                                
                                #print(combo[x])
                                #write_sheet.cell(x+2,j).value = combo[x]
                                x=x+1
                                #cnt=cnt+1
                                continue
                                                           
                            if i[0]== "部会" :
                                #print(i[1])
                                #combo.insert(y,str(combo[y])+"・部会<br/>"+str(i[1])+"<br/>")
                                combo.insert(y,str(combo[y])+str(i[1])+"<br/>")
                                #print(combo[y])
                                y=y+1
                                continue
                                
                            if i[0]== "委員・イニシアティブ名称" :
                                #print(i[1])
                                #combo.insert(z,str(combo[z])+"委員・イニシアティブ名称<br/>"+str(i[1])+"<br/>")
                                combo.insert(z,str(combo[z])+str(i[1])+"<br/>")
                                #print(combo[z])
                                write_sheet.cell(z+2,j).value = combo[z]
                                z=z+1
                                continue
                print("タグ write completed!")
                
                write_excel.save(filename=write_file)
                print("All items write completed!")
                print("--------------")
                print("Write sample data to following sheets:",write_excel.active.title,"has(after)", write_sheet.max_row, "rows,", write_sheet.max_column, "columns")
                print("=============Write sample data START=============")
                return 0

        except Exception as e:
                print("FAILED open Excel:",e)
                return -1

def close_write_excel(write_excel):
        try:
                return write_excel.save(filename=write_file) # open excel file

        except Exception as e:
                print("FAILED open Excel:",e)
                return -1
            

                   
def main():
        
        read_excel = open_read_excel()

        if read_excel != -1:
                read_1_sheet(read_excel)
        
        write_excel = open_write_excel()        
        #close_write_excel(write_excel)

main()
